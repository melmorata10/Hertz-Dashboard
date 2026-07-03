"""Agent-level AHT views for the Roadside AHT export (Agent AHT tab)."""

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_NAVY = "1D4675"
_GOLD_LIGHT = "FFF2CC"

# LineOfBusiness value in the export → site label shown on the dashboard.
# ROADSIDEPH teams are the Philippines site; all US roadside teams are Lubbock.
SITE_MAP = {
    "ROADSIDEPH": "Philippines",
    "ROADSIDE": "Lubbock, US",
    "US ROADSIDE": "Lubbock, US",
    "US RS": "Lubbock, US",
}

_REQUIRED = ["AgentName", "LineOfBusiness", "CallDate", "AHT-Inbound", "Handled", "No AHT"]


def parse_agent_aht(files) -> pd.DataFrame:
    """Read one or more Roadside AHT CSVs into tidy rows: Site / Agent / Date / Handled / HandleTime.

    Accepts a single file or a list of files; identical rows appearing in more
    than one upload (overlapping exports) are counted once. Rows flagged
    "Exclude" in the "No AHT" column carry no inbound AHT and are dropped, as
    are rows with zero handled calls.
    """
    if not isinstance(files, (list, tuple)):
        files = [files]

    frames = []
    for f in files:
        raw = pd.read_csv(f)
        raw.columns = raw.columns.str.strip()
        missing = [c for c in _REQUIRED if c not in raw.columns]
        if missing:
            raise ValueError(
                f"{getattr(f, 'name', 'File')} doesn't look like a Roadside AHT export — "
                "missing column(s): " + ", ".join(missing)
            )
        frames.append(raw)

    df = pd.concat(frames, ignore_index=True).drop_duplicates()

    df = df[df["No AHT"].astype(str).str.strip().str.lower() == "include"].copy()
    df["Handled"] = pd.to_numeric(df["Handled"], errors="coerce").fillna(0)
    df["AHT-Inbound"] = pd.to_numeric(df["AHT-Inbound"], errors="coerce")
    df = df[(df["Handled"] > 0) & df["AHT-Inbound"].notna()].copy()

    df["Date"] = pd.to_datetime(df["CallDate"], errors="coerce").dt.date
    df = df[df["Date"].notna()].copy()

    _lob = df["LineOfBusiness"].astype(str).str.strip()
    df["Site"] = _lob.str.upper().map(SITE_MAP).fillna(_lob)
    df["Agent"] = df["AgentName"].astype(str).str.strip().str.title()
    df["HandleTime"] = df["AHT-Inbound"] * df["Handled"]
    return df[["Site", "Agent", "Date", "Handled", "HandleTime"]]


def _pivot(df: pd.DataFrame, index_cols: list) -> pd.DataFrame:
    """Weighted AHT pivot: one row per index_cols combo, one column per date."""
    grp = (
        df.groupby(index_cols + ["Date"], as_index=False)
        .agg(Handled=("Handled", "sum"), HandleTime=("HandleTime", "sum"))
    )
    grp["AHT"] = grp["HandleTime"] / grp["Handled"]

    pivot = grp.pivot_table(index=index_cols, columns="Date", values="AHT")
    pivot = pivot.reindex(columns=sorted(df["Date"].unique()))
    pivot.columns = [d.strftime("%b %d") for d in pivot.columns]

    totals = df.groupby(index_cols).agg(
        Handled=("Handled", "sum"), HandleTime=("HandleTime", "sum")
    )
    pivot["Overall AHT"] = totals["HandleTime"] / totals["Handled"]
    pivot["Calls"] = totals["Handled"].astype(int)
    return pivot.round(1).reset_index()


def agent_aht_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """AHT per agent per date, grouped by site."""
    return _pivot(df, ["Site", "Agent"]).sort_values(["Site", "Agent"]).reset_index(drop=True)


def site_aht_pivot(df: pd.DataFrame) -> pd.DataFrame:
    """AHT per site per date, with an All Sites total row."""
    out = _pivot(df, ["Site"])
    total = df.copy()
    total["Site"] = "All Sites"
    return pd.concat([out, _pivot(total, ["Site"])], ignore_index=True)


def date_columns(pivot: pd.DataFrame) -> list:
    """The per-date columns of a pivot (everything except labels and totals)."""
    return [c for c in pivot.columns if c not in ("Site", "Agent", "Overall AHT", "Calls")]


def _fmt_cell(col: str, v) -> str:
    if pd.isna(v):
        return "—"
    if col == "Calls":
        return f"{int(v):,}"
    if col in ("Site", "Agent"):
        return str(v)
    return f"{float(v):,.1f}"


def aht_table_html(pivot: pd.DataFrame, title: str) -> str:
    """Rich-HTML table for pasting into Outlook / Gmail, in the dashboard's email style."""
    cols = list(pivot.columns)
    head = "".join(
        f"<th style='background:#{_NAVY};color:#ffffff;padding:6px 12px;"
        f"border:1px solid #16375c;font-size:13px;"
        f"text-align:{'left' if c in ('Site', 'Agent') else 'right'}'>{c}</th>"
        for c in cols
    )
    body = []
    for _, row in pivot.iterrows():
        is_total = str(row.iloc[0]) == "All Sites"
        weight = "700" if is_total else "400"
        bg = f"#{_GOLD_LIGHT}" if is_total else "#ffffff"
        tds = "".join(
            f"<td style='padding:5px 12px;border:1px solid #d9d9d9;font-size:13px;"
            f"text-align:{'left' if c in ('Site', 'Agent') else 'right'};"
            f"font-weight:{weight};background:{bg}'>{_fmt_cell(c, row[c])}</td>"
            for c in cols
        )
        body.append(f"<tr>{tds}</tr>")
    return (
        "<div style='font-family:Arial,sans-serif'>"
        f"<p style='font-weight:700;font-size:14px;margin:0 0 6px'>{title}</p>"
        "<table style='border-collapse:collapse'>"
        f"<tr>{head}</tr>{''.join(body)}</table>"
        "<p style='font-size:11px;color:#666;margin:6px 0 0'>"
        "AHT in seconds, weighted by handled calls</p></div>"
    )


def _write_table(ws, df: pd.DataFrame, start_row: int) -> None:
    """Write a pivot DataFrame as a styled table starting at start_row."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_NAVY)
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    total_fill = PatternFill("solid", fgColor=_GOLD_LIGHT)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        is_total = str(row.iloc[0]) == "All Sites"
        for j, col in enumerate(df.columns, start=1):
            v = row[col]
            if pd.isna(v):
                v = None
            elif col == "Calls":
                v = int(v)
            elif col not in ("Site", "Agent"):
                v = float(v)
            cell = ws.cell(row=i, column=j, value=v)
            if col == "Calls":
                cell.number_format = "#,##0"
            elif col not in ("Site", "Agent"):
                cell.number_format = "0.0"
            if is_total:
                cell.font = Font(bold=True)
                cell.fill = total_fill

    for j, col in enumerate(df.columns, start=1):
        width = max(len(str(col)) + 2, 10)
        if col in ("Site", "Agent"):
            width = max([len(str(v)) for v in df[col]] + [len(col)]) + 3
        ws.column_dimensions[get_column_letter(j)].width = width


def build_agent_aht_workbook(site_pivot: pd.DataFrame, agent_pivot: pd.DataFrame) -> bytes:
    """Two-sheet workbook: site summary with an AHT comparison chart, and the full agent view."""
    dates = date_columns(site_pivot)
    n_dates = len(dates)
    n_sites = len(site_pivot) - 1  # exclude the All Sites total row from the chart

    wb = Workbook()
    ws = wb.active
    ws.title = "AHT by Site"

    ws["A1"] = "Agent AHT Report"
    ws["A1"].font = Font(bold=True, size=14, color=_NAVY)
    ws["A2"] = "Roadside PH = Philippines · Roadside = Lubbock, US — AHT in seconds, weighted by handled calls"
    ws["A2"].font = Font(size=9, italic=True, color="666666")

    header_row = 4
    _write_table(ws, site_pivot, start_row=header_row)

    if n_dates >= 1 and n_sites >= 1:
        chart = LineChart()
        chart.title = "AHT Comparison by Site"
        chart.y_axis.title = "AHT (seconds)"
        chart.height = 9
        chart.width = 22
        data = Reference(
            ws, min_col=1, max_col=1 + n_dates,
            min_row=header_row + 1, max_row=header_row + n_sites,
        )
        chart.add_data(data, from_rows=True, titles_from_data=True)
        cats = Reference(ws, min_col=2, max_col=1 + n_dates, min_row=header_row, max_row=header_row)
        chart.set_categories(cats)
        for series in chart.series:
            series.smooth = False
        ws.add_chart(chart, f"A{header_row + len(site_pivot) + 3}")

    ws2 = wb.create_sheet("AHT per Agent")
    _write_table(ws2, agent_pivot, start_row=1)
    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = ws2.dimensions

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def report_filename(dates) -> str:
    """Agent_AHT_Report_<date range>.xlsx from the selected dates."""
    dates = sorted(dates)
    if not dates:
        return "Agent_AHT_Report.xlsx"
    if dates[0] == dates[-1]:
        return f"Agent_AHT_Report_{dates[0]:%Y-%m-%d}.xlsx"
    return f"Agent_AHT_Report_{dates[0]:%Y-%m-%d}_to_{dates[-1]:%Y-%m-%d}.xlsx"
