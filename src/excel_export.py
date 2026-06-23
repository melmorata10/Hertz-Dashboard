"""Builds the daily ASA Report workbook (.xlsx) from the dashboard's interval data.

Mirrors the layout of the existing "ASA Report" distributed to stakeholders:
one sheet per Line of Business, each with an NCO-vs-ASA chart, an ASA-per-Vendor
chart, and an auto-generated narrative describing vendor ASA patterns for the day.
"""
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import AreaChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Sheet name -> dashboard LOB name, in the exact order the stakeholder report uses.
# Derived from src/mapping.py skill entries (e.g. "...REX_Decagon" -> "Rental Extensions",
# "...CSCC_Decagon" -> "Billing/Disputes", "...TNCCSCC_Decagon" -> "TNC Billing and Dispute").
REPORT_LOB_SHEETS: dict[str, str] = {
    "Billing-Disputes": "Billing/Disputes",
    "CCM":              "CCM",
    "Roadside":         "Roadside Services",
    "Sales":            "Sales",
    "REX":              "Rental Extensions",
    "TNC Extension":    "TNC Rental Extension",
    "TNC CSCC":         "TNC Billing and Dispute",
}

_NAVY = "1D4675"
_GOLD = "FFD700"


def _fmt_time_label(hhmm: str) -> str:
    """'18:30' -> '6:30 PM'."""
    try:
        h, m = (int(p) for p in hhmm.split(":"))
    except (ValueError, AttributeError):
        return str(hhmm)
    period = "AM" if h < 12 else "PM"
    h12 = h % 12 or 12
    return f"{h12}:{m:02d} {period}"


def _lob_pivot(lob_df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """One row per Interval: NCO total, blended ASA, plus one ASA column per vendor."""
    df = lob_df.copy()
    df["_ASA_w"] = df["NCH"].fillna(0) * df["ASA"].fillna(0)
    overall = df.groupby("Interval", sort=True).agg(
        NCO=("NCO", "sum"),
        NCH=("NCH", "sum"),
        _ASA_w=("_ASA_w", "sum"),
    ).reset_index()
    safe_nch = overall["NCH"].replace(0, float("nan"))
    overall["ASA"] = (overall["_ASA_w"] / safe_nch).round(1)
    overall = overall[["Interval", "NCO", "ASA"]]

    vendors = sorted(v for v in df["Vendor"].dropna().unique() if v not in ("", "Unknown"))
    vendor_wide = df.pivot_table(index="Interval", columns="Vendor", values="ASA", aggfunc="mean")
    vendor_wide = vendor_wide.reindex(columns=vendors).reset_index()

    merged = overall.merge(vendor_wide, on="Interval", how="left").sort_values("Interval")
    return merged.reset_index(drop=True), vendors


def _vendor_narrative(lob_df: pd.DataFrame) -> str:
    """Plain-language summary of ASA patterns per vendor, in the stakeholder-report style."""
    df = lob_df[lob_df["ASA"].notna() & (lob_df["NCH"].fillna(0) > 0)].copy()
    if df.empty:
        return "No contact data available for this LOB today."

    stats = df.groupby("Vendor")["ASA"].mean().sort_values(ascending=False)
    clauses = []
    for i, vendor in enumerate(stats.index):
        v_rows = df[df["Vendor"] == vendor]
        lo, hi = v_rows["ASA"].quantile([0.1, 0.9])
        peak_row = v_rows.loc[v_rows["ASA"].idxmax()]
        descriptor = (
            "operated at elevated ASA levels for much of the day"
            if i == 0 else "stayed relatively stable"
        )
        clauses.append(
            f"{vendor} {descriptor}, generally within the {lo:.0f}–{hi:.0f}s range, "
            f"peaking at approximately {peak_row['ASA']:.0f}s around {_fmt_time_label(peak_row['Interval'])}"
        )

    sentence = clauses[0]
    for c in clauses[1:]:
        sentence += f", while {c}"
    sentence += "."

    intervals = sorted(df["Interval"].unique())
    if len(intervals) >= 4:
        tail = intervals[-2:]
        tail_mean = df[df["Interval"].isin(tail)]["ASA"].mean()
        overall_mean = df["ASA"].mean()
        if overall_mean > 0 and tail_mean < overall_mean * 0.85:
            tail_lo, tail_hi = df[df["Interval"].isin(tail)]["ASA"].quantile([0.1, 0.9])
            sentence += (
                f" ASA improved to approximately {tail_lo:.0f}–{tail_hi:.0f}s "
                f"during the final intervals."
            )
    return sentence


def _write_sheet(ws, lob_name: str, lob_df: pd.DataFrame, call_date: str):
    title_font = Font(bold=True, size=13, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_NAVY)
    narrative_font = Font(italic=True, size=11)

    ws["A1"] = f"{lob_name} — ASA Report — {call_date or 'Today'}"
    ws["A1"].font = title_font
    ws["A1"].fill = PatternFill("solid", fgColor=_NAVY)
    ws.merge_cells("A1:Q1")
    ws.row_dimensions[1].height = 22

    if lob_df.empty or lob_df["NCO"].fillna(0).sum() == 0:
        ws["A3"] = f"No {lob_name} Contact"
        ws["A3"].font = Font(italic=True, size=12)
        return

    pivot, vendors = _lob_pivot(lob_df)
    n = len(pivot)

    # ── Underlying data table (drives both charts) — placed off to the right ───
    data_col0 = 19  # column S
    headers = ["Interval", "NCO", "ASA (Blended)"] + vendors
    for j, h in enumerate(headers):
        cell = ws.cell(row=1, column=data_col0 + j, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for i, row in pivot.iterrows():
        ws.cell(row=2 + i, column=data_col0, value=row["Interval"])
        ws.cell(row=2 + i, column=data_col0 + 1, value=row["NCO"])
        ws.cell(row=2 + i, column=data_col0 + 2, value=row["ASA"])
        for k, vendor in enumerate(vendors):
            val = row.get(vendor)
            ws.cell(row=2 + i, column=data_col0 + 3 + k, value=None if pd.isna(val) else val)
    last_row = 1 + n
    last_col = data_col0 + 2 + len(vendors)
    for j in range(len(headers)):
        ws.column_dimensions[get_column_letter(data_col0 + j)].width = 12

    cats = Reference(ws, min_col=data_col0, min_row=2, max_row=last_row)

    # ── Chart 1: NCO vs ASA ──────────────────────────────────────────────────
    chart1 = AreaChart()
    chart1.title = "NCO vs ASA"
    chart1.style = 10
    chart1.y_axis.title = "Value"
    chart1.x_axis.title = "Interval"
    chart1.height, chart1.width = 11, 22
    data1 = Reference(ws, min_col=data_col0 + 1, max_col=data_col0 + 2, min_row=1, max_row=last_row)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats)
    ws.add_chart(chart1, "A3")

    # ── Chart 2: ASA Per Vendor ──────────────────────────────────────────────
    chart2 = LineChart()
    chart2.title = "ASA Per Vendor"
    chart2.style = 10
    chart2.y_axis.title = "ASA (seconds)"
    chart2.x_axis.title = "Interval"
    chart2.height, chart2.width = 11, 22
    if vendors:
        data2 = Reference(ws, min_col=data_col0 + 3, max_col=last_col, min_row=1, max_row=last_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats)
    ws.add_chart(chart2, "A26")

    # ── Narrative ─────────────────────────────────────────────────────────────
    ws["A50"] = "Analysis"
    ws["A50"].font = Font(bold=True, size=11)
    ws.merge_cells("A51:Q55")
    narrative_cell = ws["A51"]
    narrative_cell.value = _vendor_narrative(lob_df)
    narrative_cell.font = narrative_font
    narrative_cell.alignment = Alignment(wrap_text=True, vertical="top")
    for r in range(51, 56):
        ws.row_dimensions[r].height = 18


def build_asa_report_workbook(interval_df: pd.DataFrame, call_date: str = None) -> bytes:
    """Build the multi-sheet ASA Report workbook and return its bytes."""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, lob_name in REPORT_LOB_SHEETS.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        lob_df = (
            interval_df[interval_df["LOB"] == lob_name].copy()
            if interval_df is not None and not interval_df.empty
            else pd.DataFrame()
        )
        _write_sheet(ws, lob_name, lob_df, call_date)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def default_filename(call_date: str = None) -> str:
    if call_date:
        try:
            d = datetime.strptime(str(call_date), "%m/%d/%Y")
            return f"ASA_Report_{d.strftime('%Y-%m-%d')}.xlsx"
        except ValueError:
            pass
    return f"ASA_Report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
